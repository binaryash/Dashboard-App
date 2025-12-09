from core import db
from core.models import Todo
from datetime import datetime as dt
from openpyxl import load_workbook


class Todos:
    @staticmethod
    def append_to_excel(content):
        try:
            workbook = load_workbook("todos.xlsx")
            sheet = workbook.active
            next_row = sheet.max_row + 1
            current_datetime = dt.now().strftime("%Y-%m-%d %H:%M:%S")

            sheet.cell(row=next_row, column=1, value=content)
            sheet.cell(row=next_row, column=2, value=current_datetime)

            workbook.save("todos.xlsx")
        except Exception as e:
            print("Error: Unable to append content to Excel file")
            print(e)

    @staticmethod
    def create_task(content):
        new_task = Todo(content=content)
        db.session.add(new_task)
        db.session.commit()
        Todos.append_to_excel(content)
        return new_task

    @staticmethod
    def get_all_tasks():
        return Todo.query.order_by(Todo.date_created).all()

    @staticmethod
    def get_incomplete_tasks():
        return (
            Todo.query.filter_by(completion_time=None).order_by(Todo.date_created).all()
        )

    @staticmethod
    def mark_task_complete(task_id):
        task = Todo.query.get_or_404(task_id)
        if not task.completion_time:
            task.completion_time = dt.utcnow()
            task.completion_date = dt.utcnow().date()
            db.session.commit()
        return task

    @staticmethod
    def delete_task(task_id):
        task = Todo.query.get_or_404(task_id)
        db.session.delete(task)
        db.session.commit()

    @staticmethod
    def update_task(task_id, content, completion_date):
        task = Todo.query.get_or_404(task_id)
        task.content = content
        task.completion_date = (
            dt.strptime(completion_date, "%Y-%m-%d").date() if completion_date else None
        )
        db.session.commit()
        return task

    @staticmethod
    def get_task_by_id(task_id):
        return Todo.query.get_or_404(task_id)

    # In core/todo/todo.py - Add these methods to the Todos class

    @staticmethod
    def get_incomplete_tasks():
        """Get all incomplete tasks."""
        return (
            Todo.query.filter_by(completion_time=None).order_by(Todo.date_created).all()
        )

    @staticmethod
    def get_completed_tasks():
        """Get all completed tasks."""
        return (
            Todo.query.filter(Todo.completion_time.isnot(None))
            .order_by(Todo.date_created)
            .all()
        )

    @staticmethod
    def search_tasks(query):
        """Search tasks by content."""
        if not query:
            return Todos.get_all_tasks()
        search_pattern = f"%{query}%"
        return (
            Todo.query.filter(Todo.content.ilike(search_pattern))
            .order_by(Todo.date_created)
            .all()
        )
