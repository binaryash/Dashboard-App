from flask import Flask, render_template, url_for, redirect, request
from app1.feed_reader import get_feeds
from app1 import app, db
from app1.models import Todo
from app1.datetime_1 import get_current_datetime
from openpyxl import Workbook, load_workbook
from datetime import datetime as dt


@app.route('/index', methods=['POST', 'GET'])
def index():
    if request.method == 'POST':
        task_content = request.form['content']
        new_task = Todo(content=task_content)
        try:
            db.session.add(new_task)
            db.session.commit()

            # Appending the content to an excel file
            append_to_excel(task_content)
            
            return redirect('/incomplete')
        except:
            return 'There was an issue with adding your task'
    else:
        tasks = Todo.query.order_by(Todo.date_created).all()
        return render_template('todo_index.html', tasks=tasks)

def append_to_excel(content):
    try:
        # Load the Excel file
        workbook = load_workbook('todos.xlsx')
        sheet = workbook.active

        # Find the next empty row
        next_row = sheet.max_row + 1

        # Get the current date and time
        current_datetime = dt.now().strftime("%Y-%m-%d %H:%M:%S")

        sheet.cell(row=next_row, column=1, value=content)
        sheet.cell(row=next_row, column=2, value=current_datetime)

        # Save the changes
        workbook.save('todos.xlsx')
    except Exception as e:
        print('Error: Unable to append content to Excel file')
        print(e)

@app.route('/complete/<int:id>')
def complete(id):
    task = Todo.query.get_or_404(id)
    if not task.completion_time:
        task.completion_time = dt.utcnow()
        task.completion_date = dt.utcnow().date()
        db.session.commit()
    return redirect('/incomplete')




@app.route('/delete/<int:id>')
def delete(id):
    task_to_delete = Todo.query.get_or_404(id)

    try:
        db.session.delete(task_to_delete)
        db.session.commit()
        return redirect('/index')
    except:
        return 'cant delete'
    

@app.route('/update/<int:id>', methods=['GET','POST'])
def update(id):

    task = Todo.query.get_or_404(id)

    if request.method =='POST':
        task.content = request.form['content']
        task.completion_date = dt.strptime(request.form['completion_date'], '%Y-%m-%d').date() if request.form[
            'completion_date'] else None

        try:
            db.session.commit()
            return redirect('/index')
        except:
            return 'issue in updating'
    else:
        return render_template('todo_update.html', task=task)


@app.route('/incomplete')
def incomplete():
    # Get only the incomplete tasks
    tasks = Todo.query.filter_by(completion_time=None).order_by(Todo.date_created).all()
    return render_template('incomplete_tasks.html', tasks=tasks)
